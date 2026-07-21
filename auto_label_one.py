"""
自动标注脚本 - 用Qwen3-4B零样本分类器批量标注任务类型

用法：
    python auto_label.py

功能：
    读取 wildchat_2000_label.xlsx，用 Qwen3-4B 零样本分类器自动标注
    标注结果写入 Label 列，保存为 wildchat_2000_auto_labeled.xlsx
    同时保存 JSON 格式 wildchat_2000_auto_labeled.json

配置：
    修改下方 API_URL / API_KEY / MODEL 可更换分类器模型
    修改 BATCH_SAVE 控制每标注多少条自动保存一次（防中断丢数据）
"""

import json
import re
import time
import threading
import urllib.request
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===================== 配置 =====================
API_URL = "https://aigw.telecomjs.com/v1/chat/completions"
API_KEY = "sk-MrgfsKcx2YSzZhMV4ujgawqhnm2cHb7q"
MODEL = "qwen3.7-max"                      # 分类器模型
NUM_WORKERS = 40                        # 并发线程数（调大提速，调小防限流）
BATCH_SAVE = 100                       # 每标注N条自动保存一次
MAX_RETRIES = 3                        # 单条请求最大重试次数
QUERY_MAX_LEN = 500                    # 发给API的文本最大长度

INPUT_XLSX = r"D:\dqwen\one_date.xlsx"
OUTPUT_XLSX = r"D:\dqwen\qwen3.xlsx"
OUTPUT_JSON = r"D:\dqwen\qwen3.json"
LOG_FILE = r"D:\dqwen\qwen3.txt"
# =================================================

# 分类器 system prompt（来自 classify_prompt.py 的 sp_gpt）
SYSTEM_PROMPT = """# Role

You are a task classifier for an AI Gateway.

Your responsibility is to classify the entire conversation into exactly ONE of the following categories:

- simple
- complex
- coding

The classification result will be used for model routing.

Return ONLY valid JSON.

---

# Classification Rules

## 1. coding

Select "coding" if the user's primary objective is software development or source code processing.

This includes but is not limited to:

- writing code
- modifying code
- debugging
- explaining code
- reviewing code
- refactoring
- SQL
- Shell
- Python
- Java
- C++
- Go
- Rust
- JavaScript
- HTML/CSS
- Docker
- Kubernetes
- YAML
- JSON Schema
- API design
- Git
- Regular Expressions
- unit tests
- software architecture for implementation

If software engineering is the main topic, ALWAYS classify as "coding", regardless of reasoning complexity.

---

## 2. complex

Select "complex" if solving the request requires substantial reasoning, planning, analysis, synthesis, or integrating multiple constraints.

Typical characteristics:

- multi-step reasoning
- long-context understanding
- architecture or system design
- project planning
- research
- report generation
- business analysis
- legal analysis
- financial analysis
- comparing multiple solutions
- designing workflows
- generating long structured documents
- decision making with multiple factors

The answer cannot be produced by a straightforward response.

---

## 3. simple

Select "simple" if the request can be completed directly without significant reasoning or planning.

Typical examples:

- factual questions
- definitions
- translation
- grammar correction
- rewriting
- short summarization
- basic calculations
- simple recommendations
- FAQs
- short explanations

Usually requires only a direct response.

---

# Priority

If multiple categories appear applicable, use the following priority:

coding > complex > simple

---

# Conversation Scope

Classify based on the ENTIRE conversation rather than only the last user message.

If earlier turns establish that the conversation is about software development, classify as coding even if the latest message is brief.

---

# Output

Return ONLY JSON.

{
    "intent":"simple|complex|coding",
    "confidence":0.00-1.00,
    "reason":"one short sentence"
}"""


def call_api(query, retries=MAX_RETRIES):
    """调用API进行分类，返回原始输出（线程安全）
    非200状态码视为失败，自动重试，共3次机会，全失败才返回ERROR
    """
    import urllib.error
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    body = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": query[:QUERY_MAX_LEN]},
        ],
        "max_tokens": 2000,
        "enable_thinking": True,
    }
    data = json.dumps(body).encode("utf-8")

    for attempt in range(retries):
        try:
            req = urllib.request.Request(API_URL, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                return result["choices"][0]["message"]["content"].strip()
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code}"
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
        except Exception as e:
            last_err = str(e)
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
    return f"ERROR: {last_err} (重试{retries}次均失败)"


def parse_label(raw):
    """解析API输出，返回 dict: {intent, confidence, reason}"""
    result = {"intent": "UNKNOWN", "confidence": 0.0, "reason": ""}

    # 清理 thinking 标签和 markdown 代码块标记
    text = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text).strip()

    # 优先尝试整体 JSON 解析
    try:
        data = json.loads(text)
        intent_val = str(data.get("intent", "")).strip().lower()
        if intent_val in ("simple", "complex", "coding"):
            result["intent"] = intent_val.capitalize() if intent_val != "coding" else "Coding"
            result["confidence"] = float(data.get("confidence", 0.0))
            result["reason"] = str(data.get("reason", ""))
            return result
    except (json.JSONDecodeError, ValueError, AttributeError):
        pass

    # 容错：从原始文本中正则提取 intent/confidence/reason
    m_intent = re.search(r'"intent"\s*:\s*"([^"]+)"', text, re.IGNORECASE)
    m_conf = re.search(r'"confidence"\s*:\s*([0-9.]+)', text)
    m_reason = re.search(r'"reason"\s*:\s*"([^"]+)"', text)
    if m_intent:
        intent_val = m_intent.group(1).strip().lower()
        if intent_val in ("simple", "complex", "coding"):
            result["intent"] = intent_val.capitalize() if intent_val != "coding" else "Coding"
            result["confidence"] = float(m_conf.group(1)) if m_conf else 0.0
            result["reason"] = m_reason.group(1) if m_reason else ""
            return result

    # 最后回退：关键词匹配
    u = raw.upper()
    if "SIMPLE" in u:
        result["intent"] = "Simple"
        result["confidence"] = 0.5
        result["reason"] = "keyword fallback"
    elif "COMPLEX" in u:
        result["intent"] = "Complex"
        result["confidence"] = 0.5
        result["reason"] = "keyword fallback"
    elif "CODING" in u:
        result["intent"] = "Coding"
        result["confidence"] = 0.5
        result["reason"] = "keyword fallback"
    return result


def save_workbook(wb, results_json, log_f=None):
    """保存Excel、JSON和日志"""
    wb.save(OUTPUT_XLSX)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results_json, f, ensure_ascii=False, indent=2)
    if log_f:
        log_f.flush()


def process_one(row_data):
    """处理单条数据（供线程池调用），返回结果dict"""
    row_idx, no, lang, query = row_data
    raw_output = call_api(query)
    parsed = parse_label(raw_output)
    return {
        "row_idx": row_idx,
        "no": no,
        "lang": lang,
        "query": query,
        "raw_output": raw_output,
        "parsed": parsed,
    }


def main():
    print(f"读取: {INPUT_XLSX}")
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    total_rows = ws.max_row - 1  # 减去表头
    print(f"共 {total_rows} 条待标注")
    print(f"分类器: {MODEL} (零样本)")
    print(f"并发线程: {NUM_WORKERS}")
    print(f"每 {BATCH_SAVE} 条自动保存一次")
    print()

    # 添加额外输出列的表头（如果不存在）
    extra_cols = {5: "Confidence", 6: "Reason", 7: "Model"}
    for col_idx, col_name in extra_cols.items():
        header = ws.cell(row=1, column=col_idx).value
        if header != col_name:
            ws.cell(row=1, column=col_idx).value = col_name

    # 统计
    labeled_count = 0
    label_counts = Counter()
    unknown_count = 0
    error_count = 0
    results_json = []

    # 打开日志文件
    log_f = open(LOG_FILE, "w", encoding="utf-8")
    log_f.write(f"自动标注日志 | 模型={MODEL} | thinking=ON | max_tokens=2000 | 共{total_rows}条 | 并发={NUM_WORKERS}\n")
    log_f.write("=" * 100 + "\n\n")

    # ---- 第一阶段：收集已标注行 + 待标注行 ----
    skip_rows = []   # 已有标签，直接统计
    todo_rows = []    # 需要调API的行
    for row_idx in range(2, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=1).value
        if no is None:
            continue
        lang = ws.cell(row=row_idx, column=2).value or ""
        query = ws.cell(row=row_idx, column=3).value or ""

        existing_label = ws.cell(row=row_idx, column=4).value
        if existing_label and str(existing_label).strip() in ("Simple", "Complex", "Coding"):
            skip_rows.append((row_idx, no, lang, query, str(existing_label).strip()))
        else:
            todo_rows.append((row_idx, no, lang, query))

    # 处理已标注行
    for row_idx, no, lang, query, label in skip_rows:
        label_counts[label] += 1
        labeled_count += 1
        results_json.append({
            "idx": int(no),
            "language": lang,
            "query": query,
            "model": MODEL,
            "intent": label,
            "confidence": None,
            "reason": None,
            "method": "auto",
        })

    print(f"跳过已标注: {len(skip_rows)} 条，待标注: {len(todo_rows)} 条")
    if not todo_rows:
        print("全部已标注，无需处理")
        save_workbook(wb, results_json, log_f)
        log_f.close()
        return

    # ---- 第二阶段：多线程并发调API ----
    start_time = time.time()
    completed_count = 0
    lock = threading.Lock()

    def on_task_done(future):
        """回调：线程完成时更新计数"""
        nonlocal completed_count
        with lock:
            completed_count += 1

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        future_to_row = {
            executor.submit(process_one, row): row for row in todo_rows
        }

        for future in as_completed(future_to_row):
            row_info = future_to_row[future]
            try:
                result = future.result()
            except Exception as e:
                # 线程本身异常（极罕见），构造一个错误结果
                result = {
                    "row_idx": row_info[0],
                    "no": row_info[1],
                    "lang": row_info[2],
                    "query": row_info[3],
                    "raw_output": f"ERROR: thread exception: {e}",
                    "parsed": {"intent": "UNKNOWN", "confidence": 0.0, "reason": "thread exception"},
                }

            parsed = result["parsed"]
            label = parsed["intent"]
            row_idx = result["row_idx"]
            no = result["no"]
            lang = result["lang"]
            query = result["query"]
            raw_output = result["raw_output"]

            # 写入Excel（openpyxl 单线程写，as_completed 串行回主线程）
            ws.cell(row=row_idx, column=4).value = label
            ws.cell(row=row_idx, column=5).value = parsed["confidence"]
            ws.cell(row=row_idx, column=6).value = parsed["reason"]
            ws.cell(row=row_idx, column=7).value = MODEL

            # 统计
            labeled_count += 1
            if label == "UNKNOWN":
                unknown_count += 1
            if raw_output.startswith("ERROR"):
                error_count += 1
            label_counts[label] += 1

            record = {
                "idx": int(no),
                "language": lang,
                "query": query,
                "model": MODEL,
                "intent": parsed["intent"],
                "confidence": parsed["confidence"],
                "reason": parsed["reason"],
                "raw_output": raw_output,
                "method": "auto",
            }
            results_json.append(record)

            # 写日志
            log_f.write(f"[{labeled_count}/{total_rows}] idx={no} model={MODEL}\n")
            log_f.write(f"  intent={parsed['intent']}  confidence={parsed['confidence']:.2f}  reason={parsed['reason']}\n")
            log_f.write(f"  query: {query[:200]}\n")
            log_f.write("  " + "-" * 100 + "\n")
            log_f.flush()

            # 进度显示
            with lock:
                done = completed_count
            if done % 50 == 0 or done == len(todo_rows):
                elapsed = time.time() - start_time
                speed = done / elapsed if elapsed > 0 else 0
                remaining = len(todo_rows) - done
                eta = remaining / speed if speed > 0 else 0
                print(
                    f"  [{labeled_count}/{total_rows}] "
                    f"Simple={label_counts.get('Simple',0)} Complex={label_counts.get('Complex',0)} "
                    f"Coding={label_counts.get('Coding',0)} Unknown={unknown_count} "
                    f"| {speed:.1f}条/s ETA={eta/60:.1f}min"
                )

            # 定期保存
            if done % BATCH_SAVE == 0:
                save_workbook(wb, results_json, log_f)
                print(f"  >> 已自动保存 ({labeled_count}/{total_rows})")

    # 最终保存
    save_workbook(wb, results_json, log_f)

    elapsed = time.time() - start_time
    # 写日志汇总
    log_f.write("\n" + "=" * 100 + "\n")
    log_f.write("标注汇总\n")
    log_f.write("=" * 100 + "\n")
    log_f.write(f"总标注: {labeled_count} 条\n")
    log_f.write(f"模型: {MODEL}\n")
    log_f.write(f"耗时: {elapsed/60:.1f} 分钟\n")
    log_f.write(f"错误: {error_count} 条\n")
    log_f.write(f"UNKNOWN: {unknown_count} 条\n")
    log_f.write(f"\n标签分布:\n")
    for k in ["Simple", "Complex", "Coding", "UNKNOWN"]:
        v = label_counts.get(k, 0)
        if v > 0:
            log_f.write(f"  {k}: {v} ({v/labeled_count*100:.1f}%)\n")
    if unknown_count > 0:
        unknowns = [r for r in results_json if r["intent"] == "UNKNOWN"]
        log_f.write(f"\nUNKNOWN样本 ({len(unknowns)}条):\n")
        for u in unknowns[:20]:
            log_f.write(f"  idx={u['idx']} raw={u.get('raw_output','')[:100]} query={u['query'][:80]}\n")
    log_f.close()
    print(f"\n{'='*60}")
    print(f"标注完成")
    print(f"{'='*60}")
    print(f"  总标注: {labeled_count} 条")
    print(f"  耗时: {elapsed/60:.1f} 分钟")
    print(f"  错误: {error_count} 条 (API调用失败)")
    print(f"  UNKNOWN: {unknown_count} 条 (模型输出无法解析)")
    print(f"\n  标签分布:")
    for k in ["Simple", "Complex", "Coding", "UNKNOWN"]:
        v = label_counts.get(k, 0)
        if v > 0:
            print(f"    {k}: {v} ({v/labeled_count*100:.1f}%)")

    # 按语言统计
    zh = [r for r in results_json if r["language"] == "Chinese"]
    en = [r for r in results_json if r["language"] == "English"]
    print(f"\n  按语言:")
    for lang_name, lang_items in [("Chinese", zh), ("English", en)]:
        lc = Counter(r["intent"] for r in lang_items)
        print(f"    {lang_name}({len(lang_items)}): Simple={lc['Simple']} Complex={lc['Complex']} Coding={lc['Coding']}")

    print(f"\n  Excel: {OUTPUT_XLSX}")
    print(f"  JSON:  {OUTPUT_JSON}")

    # UNKNOWN样本提示
    if unknown_count > 0:
        unknowns = [r for r in results_json if r["intent"] == "UNKNOWN"]
        print(f"\n  ⚠️ UNKNOWN样本 ({len(unknowns)}条):")
        for u in unknowns[:10]:
            print(f"    idx={u['idx']} raw={u.get('raw_output','')[:60]} query={u['query'][:60]}...")
        if len(unknowns) > 10:
            print(f"    ...还有{len(unknowns)-10}条，详见JSON文件")


if __name__ == "__main__":
    main()
