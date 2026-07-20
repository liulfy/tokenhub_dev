"""
自动标注脚本 - 用Qwen3-4B零样本分类器批量标注任务类型

用法：
    python auto_label.py

功能：
    读取 wildchat_2000_label.xlsx，用 Qwen3-4B 零样本分类器自动标注
    标注结果写入 Label 列，保存为 wildchat_2000_auto_labeled.xlsx
    同时保存 JSON 格式 wildchat_2000_auto_labeled.json

配置：
    修改下方 API_URL / API_KEY / MODEL 可更换分类器模型
    修改 BATCH_SAVE 控制每标注多少条自动保存一次（防中断丢数据）
"""

import json
import time
import urllib.request
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===================== 配置 =====================
API_URL = "https://wishub-x6.ctyun.cn/v1/chat/completions"
API_KEY = "0ef2d0dbdf334d48b81d528ffd127d3c"
MODEL = "Qwen3-4B"                    # 分类器模型
BATCH_SAVE = 100                       # 每标注N条自动保存一次
REQUEST_INTERVAL = 0.3                 # API请求间隔(秒)，防限流
MAX_RETRIES = 3                        # 单条请求最大重试次数
QUERY_MAX_LEN = 500                    # 发给API的文本最大长度

INPUT_XLSX = r"D:\LLM_Routing_Datasets\wildchat_2000_label.xlsx"
OUTPUT_XLSX = r"D:\LLM_Routing_Datasets\wildchat_2000_auto_labeled.xlsx"
OUTPUT_JSON = r"D:\LLM_Routing_Datasets\wildchat_2000_auto_labeled.json"
# =================================================

# 分类器 system prompt
SYSTEM_PROMPT = (
    "你是任务路由分类器。根据用户请求判断任务类型：\n"
    "- Simple：简单问答、翻译、常识查询、简单计算、列举、打招呼、简短询问\n"
    "- Complex：深度分析、多角度论述、长文写作、创意写作、系统设计、跨学科推理\n"
    "- Coding：写代码、debug、代码审查、编程问题、技术实现\n"
    "只输出一个词：Simple、Complex 或 Coding。"
)


def call_api(query, retries=MAX_RETRIES):
    """调用API进行分类，返回原始输出"""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    body = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": query[:QUERY_MAX_LEN]},
        ],
        "max_tokens": 20,
        "enable_thinking": False,
    }
    data = json.dumps(body).encode("utf-8")

    for attempt in range(retries):
        try:
            req = urllib.request.Request(API_URL, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                return result["choices"][0]["message"]["content"].strip()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))  # 递增等待
            else:
                return f"ERROR: {e}"


def parse_label(raw):
    """解析API输出为标准标签"""
    u = raw.upper()
    if "SIMPLE" in u:
        return "Simple"
    if "COMPLEX" in u:
        return "Complex"
    if "CODING" in u:
        return "Coding"
    return "UNKNOWN"


def save_workbook(wb, results_json):
    """保存Excel和JSON"""
    wb.save(OUTPUT_XLSX)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results_json, f, ensure_ascii=False, indent=2)


def main():
    print(f"读取: {INPUT_XLSX}")
    wb = load_workbook(INPUT_XLSX)
    ws = wb.active

    total_rows = ws.max_row - 1  # 减去表头
    print(f"共 {total_rows} 条待标注")
    print(f"分类器: {MODEL} (零样本)")
    print(f"每 {BATCH_SAVE} 条自动保存一次")
    print()

    # 统计
    labeled_count = 0
    label_counts = Counter()
    unknown_count = 0
    error_count = 0
    results_json = []

    start_time = time.time()

    for row_idx in range(2, ws.max_row + 1):
        no = ws.cell(row=row_idx, column=1).value
        if no is None:
            continue

        lang = ws.cell(row=row_idx, column=2).value or ""
        query = ws.cell(row=row_idx, column=3).value or ""

        # 跳过已有标签的行（支持断点续标）
        existing_label = ws.cell(row=row_idx, column=4).value
        if existing_label and str(existing_label).strip() in ("Simple", "Complex", "Coding"):
            label_counts[str(existing_label).strip()] += 1
            labeled_count += 1
            results_json.append({
                "idx": int(no),
                "language": lang,
                "query": query,
                "label": str(existing_label).strip(),
                "method": "auto",
            })
            continue

        # 调用API分类
        raw_output = call_api(query)
        label = parse_label(raw_output)

        # 写入Excel
        ws.cell(row=row_idx, column=4).value = label

        # 统计
        labeled_count += 1
        if label == "UNKNOWN":
            unknown_count += 1
        if raw_output.startswith("ERROR"):
            error_count += 1
        label_counts[label] += 1

        results_json.append({
            "idx": int(no),
            "language": lang,
            "query": query,
            "label": label,
            "raw_output": raw_output,
            "method": "auto",
        })

        # 进度显示
        if labeled_count % 50 == 0:
            elapsed = time.time() - start_time
            speed = labeled_count / elapsed if elapsed > 0 else 0
            eta = (total_rows - labeled_count) / speed if speed > 0 else 0
            print(
                f"  [{labeled_count}/{total_rows}] "
                f"Simple={label_counts['Simple']} Complex={label_counts['Complex']} "
                f"Coding={label_counts['Coding']} Unknown={unknown_count} "
                f"| {speed:.1f}条/s ETA={eta/60:.1f}min"
            )

        # 定期保存
        if labeled_count % BATCH_SAVE == 0:
            save_workbook(wb, results_json)
            print(f"  >> 已自动保存 ({labeled_count}/{total_rows})")

        time.sleep(REQUEST_INTERVAL)

    # 最终保存
    save_workbook(wb, results_json)

    elapsed = time.time() - start_time
    print(f"\n{'='*60}")
    print(f"标注完成")
    print(f"{'='*60}")
    print(f"  总标注: {labeled_count} 条")
    print(f"  耗时: {elapsed/60:.1f} 分钟")
    print(f"  错误: {error_count} 条 (API调用失败)")
    print(f"  UNKNOWN: {unknown_count} 条 (模型输出无法解析)")
    print(f"\n  标签分布:")
    for k in ["Simple", "Complex", "Coding", "UNKNOWN"]:
        v = label_counts.get(k, 0)
        if v > 0:
            print(f"    {k}: {v} ({v/labeled_count*100:.1f}%)")

    # 按语言统计
    zh = [r for r in results_json if r["language"] == "Chinese"]
    en = [r for r in results_json if r["language"] == "English"]
    print(f"\n  按语言:")
    for lang_name, lang_items in [("Chinese", zh), ("English", en)]:
        lc = Counter(r["label"] for r in lang_items)
        print(f"    {lang_name}({len(lang_items)}): Simple={lc['Simple']} Complex={lc['Complex']} Coding={lc['Coding']}")

    print(f"\n  Excel: {OUTPUT_XLSX}")
    print(f"  JSON:  {OUTPUT_JSON}")

    # UNKNOWN样本提示
    if unknown_count > 0:
        unknowns = [r for r in results_json if r["label"] == "UNKNOWN"]
        print(f"\n  ⚠️ UNKNOWN样本 ({len(unknowns)}条):")
        for u in unknowns[:10]:
            print(f"    idx={u['idx']} raw={u.get('raw_output','')[:60]} query={u['query'][:60]}...")
        if len(unknowns) > 10:
            print(f"    ...还有{len(unknowns)-10}条，详见JSON文件")


if __name__ == "__main__":
    main()
