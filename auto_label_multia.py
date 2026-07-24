"""
单模型多轮对话自动标注脚本

调用 deepseek-v4-pro 对多轮对话进行 Simple/Complex/Coding 分类。

关键设计：将完整对话放入单个 user 消息中，避免模型直接续写对话。

输出文件（保存在 OUTPUT_DIR 下）：
    deepseek_multi_turn.xlsx / .json

用法：
    python auto_label_multi.py
"""

import json
import os
import re
import time
import threading
import urllib.request
import urllib.error
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook

# ===================== 配置 =====================
API_URL =  "https://wishub-x6.ctyun.cn/v1/chat/completions"
API_KEY = "0ef2d0dbdf334d48b81d528ffd127d3c"
MODEL = "Qwen3-4B"
MODEL_SHORT = "qwen3-4"

NUM_WORKERS = 2                         # 并发线程数
BATCH_SAVE = 50                         # 每完成N条自动保存一次
MAX_RETRIES = 3                         # 单条请求最大重试次数
TURN_MAX_LEN = 2000                      # 单轮最大字符数（截断用）
MAX_CONTEXT_CHARS = 6000                # 对话上下文最大总字符数

INPUT_XLSX = r"C:\Users\27942\Desktop\多模型打标\mult_date.xlsx"
OUTPUT_DIR = r"C:\Users\27942\Desktop\多模型打标"
LOG_FILE = r"C:\Users\27942\Desktop\多模型打标\qwen3-4.txt"
# =================================================

# 分类器 system prompt（简洁明确）
SYSTEM_PROMPT = """You are a task classifier. You MUST respond with ONLY a JSON object, no other text.

Categories:
- simple: factual questions, definitions, translation, grammar, rewriting, short summarization, basic calculations, simple recommendations, FAQs, short explanations
- complex: multi-step reasoning, long-context understanding, architecture/system design, project planning, research, report generation, business/legal/financial analysis, comparing multiple solutions, designing workflows, generating long structured documents, decision making with multiple factors
- coding: software development, writing/modifying/debugging/explaining/reviewing code, SQL, Shell, Python, Java, C++, Go, Rust, JavaScript, HTML/CSS, Docker, Kubernetes, YAML, API design, Git, unit tests

Priority: coding > complex > simple

Multi-turn rules:
- If earlier turns establish the topic is coding, classify as coding even if the last message is brief
- If earlier turns involve complex analysis and the last message continues that work, classify as complex
- A brief follow-up like "give me another example" or "translate this" is simple even in a long conversation

Respond with EXACTLY this JSON format and nothing else:
{"intent":"simple","confidence":0.95,"reason":"brief explanation"}"""


# ============ 解析与截断 ============

def parse_conversation(conv_str):
    """解析 Conversation 字符串为多轮消息列表"""
    segments = re.split(r'(?=User:\s|Assistant:\s)', conv_str)
    raw_msgs = []
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        if seg.startswith("User:"):
            content = seg[5:].strip()
            if content:
                raw_msgs.append({"role": "user", "content": content})
        elif seg.startswith("Assistant:"):
            content = seg[10:].strip()
            if content:
                raw_msgs.append({"role": "assistant", "content": content})

    messages = []
    for msg in raw_msgs:
        if messages and messages[-1]["role"] == msg["role"]:
            messages[-1]["content"] += "\n" + msg["content"]
        else:
            messages.append({"role": msg["role"], "content": msg["content"]})
    return messages


def format_conversation_text(messages):
    """将消息列表格式化为文本，截断超长部分"""
    parts = []
    total_chars = 0
    for msg in messages:
        role = "User" if msg["role"] == "user" else "Assistant"
        content = msg["content"]
        # 单轮截断
        if len(content) > TURN_MAX_LEN:
            half = TURN_MAX_LEN // 2 - 20
            content = content[:half] + f"\n...[truncated {len(content)-TURN_MAX_LEN} chars]...\n" + content[-half:]
        parts.append(f"{role}: {content}")
        total_chars += len(content) + 10
        if total_chars >= MAX_CONTEXT_CHARS:
            break
    return "\n\n".join(parts)


# ============ API 调用与解析 ============

def call_api(conv_text, retries=MAX_RETRIES):
    """调用API，将对话文本放入user消息中"""
    user_msg = f"""Below is a multi-turn conversation. Classify the LAST user message into one of: simple, complex, or coding.

Conversation:
{conv_text}

Respond with ONLY JSON: {{\"intent\":\"simple|complex|coding\",\"confidence\":0.0-1.0,\"reason\":\"brief explanation\"}}"""

    api_messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_msg},
    ]

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }
    body = {
        "model": MODEL,
        "messages": api_messages,
        "max_tokens": 2000,
        "temperature": 0.1,
    }
    data = json.dumps(body, ensure_ascii=False).encode("utf-8")

    last_err = "unknown"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(API_URL, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=120) as resp:
                result = json.loads(resp.read().decode("utf-8"))
                msg = result["choices"][0]["message"]
                content = msg.get("content", "") or ""
                reasoning = msg.get("reasoning_content", "") or ""
                # 如果content为空，拼接reasoning用于后续解析
                if not content and reasoning:
                    content = reasoning
                # 如果content和reasoning都有，合并（reasoning在前）
                elif content and reasoning:
                    content = reasoning + "\n" + content
                return content.strip()
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code}"
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
        except Exception as e:
            last_err = str(e)
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
                continue
    return f"ERROR: {last_err} (retried {retries} times)"


def parse_label(raw):
    """解析API输出，返回 dict: {intent, confidence, reason}"""
    result = {"intent": "UNKNOWN", "confidence": 0.0, "reason": ""}

    # 去除 think 标签
    THINK_START = "\u003c\u0074\u0068\u0069\u006e\u006b\u003e"
    THINK_END = "\u003c\u002f\u0074\u0068\u0069\u006e\u006b\u003e"
    text = re.sub(rf"{THINK_START}.*?{THINK_END}", "", raw, flags=re.DOTALL).strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text).strip()

    # 处理空响应
    if not text:
        return result

    # JSON 解析
    try:
        data = json.loads(text)
        intent_val = str(data.get("intent", "")).strip().lower()
        if intent_val in ("simple", "complex", "coding"):
            result["intent"] = intent_val.capitalize() if intent_val != "coding" else "Coding"
            result["confidence"] = float(data.get("confidence", 0.0))
            result["reason"] = str(data.get("reason", ""))
            return result
    except (json.JSONDecodeError, ValueError, AttributeError):
        pass

    # 从文本中提取最后一个JSON块（处理reasoning_content+content合并的情况）
    json_blocks = re.findall(r'\{[^{}]*"intent"[^{}]*\}', text)
    if json_blocks:
        try:
            data = json.loads(json_blocks[-1])
            intent_val = str(data.get("intent", "")).strip().lower()
            if intent_val in ("simple", "complex", "coding"):
                result["intent"] = intent_val.capitalize() if intent_val != "coding" else "Coding"
                result["confidence"] = float(data.get("confidence", 0.0))
                result["reason"] = str(data.get("reason", ""))
                return result
        except (json.JSONDecodeError, ValueError, AttributeError):
            pass

    # 正则兜底
    m_intent = re.search(r'"intent"\s*:\s*"([^"]+)"', text, re.IGNORECASE)
    m_conf = re.search(r'"confidence"\s*:\s*([0-9.]+)', text)
    m_reason = re.search(r'"reason"\s*:\s*"([^"]+)"', text)
    if m_intent:
        intent_val = m_intent.group(1).strip().lower()
        if intent_val in ("simple", "complex", "coding"):
            result["intent"] = intent_val.capitalize() if intent_val != "coding" else "Coding"
            result["confidence"] = float(m_conf.group(1)) if m_conf else 0.0
            result["reason"] = m_reason.group(1) if m_reason else ""
            return result

    # 关键词兜底
    u = raw.upper()
    if "SIMPLE" in u:
        result["intent"] = "Simple"
        result["confidence"] = 0.5
        result["reason"] = "keyword fallback"
    elif "COMPLEX" in u:
        result["intent"] = "Complex"
        result["confidence"] = 0.5
        result["reason"] = "keyword fallback"
    elif "CODING" in u:
        result["intent"] = "Coding"
        result["confidence"] = 0.5
        result["reason"] = "keyword fallback"
    return result


# ============ 单条数据处理 ============

def process_one_row(row_data):
    """处理单条数据：解析对话 -> 构造请求 -> 调用API -> 解析结果"""
    row_idx, no, lang, turn, conv_str = row_data

    messages = parse_conversation(conv_str)
    if not messages:
        return {
            "row_idx": row_idx, "no": no, "lang": lang, "turn": turn,
            "parsed": {"intent": "UNKNOWN", "confidence": 0.0, "reason": "parse failure"},
            "raw": "ERROR: parse failure", "latency": 0.0,
        }

    conv_text = format_conversation_text(messages)

    t_start = time.time()
    raw_output = call_api(conv_text)
    latency = time.time() - t_start
    parsed = parse_label(raw_output)

    # 如果结果为UNKNOWN且不是parse failure，重试一次
    if parsed["intent"] == "UNKNOWN" and not raw_output.startswith("ERROR") and latency < 5:
        t_start = time.time()
        raw_output = call_api(conv_text)
        latency = time.time() - t_start
        parsed = parse_label(raw_output)

    return {
        "row_idx": row_idx, "no": no, "lang": lang, "turn": turn,
        "parsed": parsed, "raw": raw_output, "latency": latency,
    }


# ============ 主流程 ============

def main():
    print(f"读取: {INPUT_XLSX}")
    src_wb = load_workbook(INPUT_XLSX, data_only=True)
    src_ws = src_wb.active

    # 读取所有行数据
    rows_data = []
    for row_idx in range(2, src_ws.max_row + 1):
        no = src_ws.cell(row=row_idx, column=1).value
        if no is None:
            break
        lang = src_ws.cell(row=row_idx, column=2).value or ""
        turn = src_ws.cell(row=row_idx, column=3).value or 0
        conv = src_ws.cell(row=row_idx, column=4).value or ""
        rows_data.append((row_idx, no, lang, turn, str(conv)))
    src_wb.close()

    total = len(rows_data)
    print(f"共 {total} 条数据")
    print(f"模型: {MODEL}")
    print(f"并发线程: {NUM_WORKERS}")
    print(f"每 {BATCH_SAVE} 条自动保存一次")
    print()

    # 创建输出 Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"
    headers = ["No", "Language", "Turn", "Label", "Confidence", "Reason", "Latency(s)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h

    # 数据存储
    json_records = []
    row_num = 2  # 从第2行开始写

    # 统计
    label_counts = Counter()
    unknown_count = 0
    error_count = 0
    lock = threading.Lock()
    completed = [0]

    log_f = open(LOG_FILE, "w", encoding="utf-8")
    log_f.write(f"单模型多轮对话标注日志 | model={MODEL} | total={total} | workers={NUM_WORKERS}\n")
    log_f.write("=" * 120 + "\n\n")

    def save_all():
        """保存 Excel 和 JSON"""
        xlsx_path = os.path.join(OUTPUT_DIR, f"{MODEL_SHORT}_multi_turn.xlsx")
        json_path = os.path.join(OUTPUT_DIR, f"{MODEL_SHORT}_multi_turn.json")
        wb.save(xlsx_path)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_records, f, ensure_ascii=False, indent=2)

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        future_to_row = {
            executor.submit(process_one_row, row): row for row in rows_data
        }

        for future in as_completed(future_to_row):
            row_info = future_to_row[future]
            try:
                result = future.result()
            except Exception as e:
                result = {
                    "row_idx": row_info[0], "no": row_info[1],
                    "lang": row_info[2], "turn": row_info[3],
                    "parsed": {"intent": "UNKNOWN", "confidence": 0.0, "reason": f"exception: {e}"},
                    "raw": f"ERROR: {e}", "latency": 0.0,
                }

            no = result["no"]
            parsed = result["parsed"]
            label = parsed["intent"]
            latency = result["latency"]

            # 写入 Excel
            ws.cell(row=row_num, column=1).value = no
            ws.cell(row=row_num, column=2).value = result["lang"]
            ws.cell(row=row_num, column=3).value = result["turn"]
            ws.cell(row=row_num, column=4).value = label
            ws.cell(row=row_num, column=5).value = parsed["confidence"]
            ws.cell(row=row_num, column=6).value = parsed["reason"]
            ws.cell(row=row_num, column=7).value = round(latency, 3)
            row_num += 1

            # 统计
            label_counts[label] += 1
            if label == "UNKNOWN":
                unknown_count += 1
            if result["raw"].startswith("ERROR"):
                error_count += 1

            # JSON 记录
            json_records.append({
                "idx": int(no),
                "language": result["lang"],
                "turn": result["turn"],
                "intent": label,
                "confidence": parsed["confidence"],
                "reason": parsed["reason"],
                "latency": round(latency, 3),
                "model": MODEL,
            })

            # 日志
            with lock:
                completed[0] += 1
                done = completed[0]
                log_f.write("[%d/%d] idx=%s turn=%s label=%s conf=%.2f lat=%.2fs\n" % (
                    done, total, no, result["turn"], label, parsed["confidence"], latency))
                log_f.flush()

            # 进度
            if done % 50 == 0 or done == total:
                elapsed = time.time() - start_time
                speed = done / elapsed if elapsed > 0 else 0
                remaining = total - done
                eta = remaining / speed if speed > 0 else 0
                label_str = " ".join("%s=%d" % (k, v) for k, v in label_counts.most_common())
                print("  [%d/%d] %s | %.1f条/s ETA=%.1fmin" % (done, total, label_str, speed, eta/60))

            # 定期保存
            if done % BATCH_SAVE == 0:
                save_all()
                log_f.flush()
                print("  >> 已自动保存 (%d/%d)" % (done, total))

    # 最终保存
    save_all()
    elapsed = time.time() - start_time

    # 日志汇总
    log_f.write("\n" + "=" * 120 + "\n")
    log_f.write("标注汇总\n")
    log_f.write("=" * 120 + "\n")
    log_f.write("总标注: %d 条\n" % total)
    log_f.write("耗时: %.1f 分钟\n\n" % (elapsed / 60))

    log_f.write("--- %s ---\n" % MODEL)
    t = sum(label_counts.values())
    for k in ["Simple", "Complex", "Coding", "UNKNOWN"]:
        v = label_counts.get(k, 0)
        if v > 0:
            log_f.write("  %s: %d (%.1f%%)\n" % (k, v, v/t*100))
    log_f.write("  错误: %d 条\n" % error_count)
    log_f.write("  UNKNOWN: %d 条\n" % unknown_count)
    log_f.close()

    # 控制台输出
    print("\n" + "=" * 70)
    print("单模型多轮对话标注完成")
    print("=" * 70)
    print("  模型: %s" % MODEL)
    print("  总标注: %d 条" % total)
    print("  耗时: %.1f 分钟\n" % (elapsed / 60))

    t = sum(label_counts.values())
    s = " ".join("%s=%d(%d%%)" % (k, label_counts.get(k,0), label_counts.get(k,0)/t*100)
                 for k in ["Simple", "Complex", "Coding"])
    err = " err=%d" % error_count if error_count > 0 else ""
    unk = " unk=%d" % unknown_count if unknown_count > 0 else ""
    print("  标签分布: %s%s%s" % (s, err, unk))

    print("\n  输出文件:")
    print("    %s_multi_turn.xlsx / .json" % MODEL_SHORT)
    print("    日志: %s" % os.path.basename(LOG_FILE))


if __name__ == "__main__":
    main()
